"""Audit-friendly XLSX export for a normalized ClaimGuard case result."""

from __future__ import annotations

import json
import os
from collections import OrderedDict
from collections.abc import Mapping, Sequence
from datetime import UTC, date, datetime
from decimal import Decimal
from enum import Enum
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.exports.common import (
    approved_challenge_lines,
    case_record,
    collection,
    compute_financial_summary,
    first_value,
    json_safe,
    liability_record,
    money,
)

SHEET_ORDER = (
    "Claim",
    "Liability",
    "Invoices",
    "Pages",
    "Lines",
    "Checks",
    "Mappings",
    "Comparisons",
    "Challenges",
    "Evidence",
    "Versions",
    "Audit",
)

NAVY = "17324D"
BLUE = "2E74B5"
PALE_BLUE = "E8EEF5"
PALE_RED = "FDECEC"
PALE_GREEN = "EAF5EE"
MUTED = "5F6B76"
WHITE = "FFFFFF"
GRID = "D8DEE6"
THIN = Side(style="thin", color=GRID)


DEFAULT_COLUMNS: dict[str, tuple[str, ...]] = {
    "Liability": ("status", "human_confirmed", "confirmed_by", "confirmed_at"),
    "Invoices": (
        "id",
        "invoice_number",
        "invoice_date",
        "repairer",
        "vehicle_registration",
        "net_total",
        "vat_total",
        "non_vat_total",
        "gross_total",
    ),
    "Pages": ("document_id", "page_number", "classification", "confidence"),
    "Lines": (
        "id",
        "invoice_id",
        "description",
        "part_number",
        "quantity",
        "unit",
        "unit_price_net",
        "invoice_net",
        "vat_rate",
        "page_number",
    ),
    "Checks": ("id", "line_id", "check_type", "status", "severity", "message"),
    "Mappings": (
        "line_id",
        "status",
        "ontology_item_id",
        "mapping_confidence",
        "rationale",
    ),
    "Comparisons": (
        "line_id",
        "invoice_net",
        "ontology_price_net",
        "historical_median_net",
        "historical_count",
        "challenge_price_net",
        "challenge_amount_net",
    ),
    "Challenges": (
        "line_id",
        "description",
        "approved",
        "challengeable",
        "included_in_letter",
        "invoice_net",
        "challenge_price_net",
        "computed_challenge_amount_net",
        "computed_vat_impact",
        "computed_gross_effect",
        "reason",
    ),
    "Evidence": ("id", "type", "source", "captured_at", "summary", "url"),
    "Versions": ("type", "version", "status", "effective_at", "source"),
    "Audit": ("id", "timestamp", "actor", "action", "entity_type", "entity_id", "note"),
}


def _flatten(record: Mapping[str, Any], prefix: str = "") -> OrderedDict[str, Any]:
    flattened: OrderedDict[str, Any] = OrderedDict()
    for key, value in record.items():
        name = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, Mapping):
            nested = _flatten(value, name)
            if nested:
                flattened.update(nested)
            else:
                flattened[name] = "{}"
        elif isinstance(value, (list, tuple, set)):
            flattened[name] = json.dumps(json_safe(value), ensure_ascii=False, sort_keys=True)
        else:
            flattened[name] = value
    return flattened


def _nested_records(
    result: Mapping[str, Any],
    top_keys: Sequence[str],
    parent_keys: Sequence[str],
    nested_keys: Sequence[str],
    parent_id_keys: Sequence[str],
    parent_label: str,
) -> list[dict[str, Any]]:
    direct = collection(result, top_keys)
    if direct:
        return direct
    records: list[dict[str, Any]] = []
    for parent in collection(result, parent_keys):
        parent_id = first_value(parent, parent_id_keys, "")
        for nested_key in nested_keys:
            nested = parent.get(nested_key)
            if isinstance(nested, Sequence) and not isinstance(nested, (str, bytes, bytearray)):
                for item in nested:
                    if isinstance(item, Mapping):
                        records.append({parent_label: parent_id, **dict(item)})
                break
    return records


def _claim_rows(result: Mapping[str, Any]) -> list[dict[str, Any]]:
    claim = _flatten(case_record(result))
    summary = compute_financial_summary(result)
    rows = [{"section": "Claim", "field": key, "value": value} for key, value in claim.items()]
    rows.extend(
        {"section": "Computed summary", "field": key, "value": value}
        for key, value in {
            "invoice_price_net": summary.invoice_price_net,
            "challenge_price_net": summary.challenge_price_net,
            "challenge_amount_net": summary.challenge_amount_net,
            "vat_impact": summary.vat_impact,
            "gross_effect": summary.gross_effect,
            "mot_non_vat_total": summary.mot_non_vat_total,
            "challenge_percentage": summary.challenge_percentage,
            "challenged_line_count": summary.challenged_line_count,
        }.items()
    )
    return rows


def _challenge_rows(result: Mapping[str, Any]) -> list[dict[str, Any]]:
    raw = collection(result, ("challenges", "challenge_lines"))
    if not raw:
        raw = collection(result, ("comparisons", "comparison_results"))
    included = {line.line_id: line for line in approved_challenge_lines(result)}
    rows: list[dict[str, Any]] = []
    for record in raw:
        row = dict(record)
        line_id = str(first_value(row, ("line_id", "invoice_line_id", "id"), ""))
        computed = included.get(line_id)
        row["included_in_letter"] = computed is not None
        if computed is not None:
            row["computed_challenge_amount_net"] = computed.challenge_amount_net
            row["computed_vat_impact"] = computed.vat_impact
            row["computed_gross_effect"] = computed.gross_effect
        rows.append(row)
    return rows


def _comparison_rows(result: Mapping[str, Any]) -> list[dict[str, Any]]:
    comparisons = collection(result, ("comparisons", "comparison_results"))
    challenges = {
        str(first_value(row, ("line_id", "invoice_line_id"), "")): row
        for row in collection(result, ("challenges", "challenge_lines"))
    }
    lines = {
        str(first_value(row, ("id", "line_id", "invoice_line_id"), "")): row
        for row in collection(result, ("lines", "invoice_lines", "line_items"))
    }
    rows: list[dict[str, Any]] = []
    for comparison in comparisons:
        row = dict(comparison)
        line_id = str(first_value(row, ("line_id", "invoice_line_id", "id"), ""))
        challenge = challenges.get(line_id)
        if challenge is not None:
            line = lines.get(line_id, {})
            invoice_net = money(
                first_value(
                    challenge,
                    ("invoice_net",),
                    first_value(row, ("invoice_net",), first_value(line, ("invoice_net",), 0)),
                )
            )
            challenge_price = money(first_value(challenge, ("challenge_price_net",), invoice_net))
            challenge_amount = max(invoice_net - challenge_price, Decimal("0.00"))
            if challenge.get("challengeable") is False:
                challenge_price = invoice_net
                challenge_amount = Decimal("0.00")
            row["invoice_net"] = invoice_net
            row["challenge_price_net"] = challenge_price
            row["challenge_amount_net"] = challenge_amount
        rows.append(row)
    return rows


def _sheet_records(result: Mapping[str, Any], sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name == "Claim":
        return _claim_rows(result)
    if sheet_name == "Liability":
        liability = liability_record(result)
        return [liability] if liability else []
    if sheet_name == "Invoices":
        return collection(result, ("invoices", "invoice_units"))
    if sheet_name == "Pages":
        return _nested_records(
            result,
            ("pages", "document_pages"),
            ("documents",),
            ("pages", "document_pages"),
            ("id", "document_id", "file_name"),
            "document_id",
        )
    if sheet_name == "Lines":
        return _nested_records(
            result,
            ("lines", "invoice_lines", "line_items"),
            ("invoices", "invoice_units"),
            ("lines", "invoice_lines", "line_items"),
            ("id", "invoice_id", "invoice_number"),
            "invoice_id",
        )
    if sheet_name == "Checks":
        return collection(result, ("checks", "calculation_checks", "findings"))
    if sheet_name == "Mappings":
        return collection(result, ("mappings", "mapping_results"))
    if sheet_name == "Comparisons":
        return _comparison_rows(result)
    if sheet_name == "Challenges":
        return _challenge_rows(result)
    if sheet_name == "Evidence":
        records = collection(result, ("evidence", "research_evidence"))
        if records:
            return records
        liability = liability_record(result)
        nested = liability.get("evidence")
        return (
            [dict(item) for item in nested if isinstance(item, Mapping)]
            if isinstance(nested, Sequence) and not isinstance(nested, (str, bytes, bytearray))
            else []
        )
    if sheet_name == "Versions":
        return collection(
            result,
            ("versions", "report_versions", "ontology_versions", "policy_versions"),
        )
    if sheet_name == "Audit":
        return collection(result, ("audit", "audit_log", "audit_events"))
    return []


def _column_names(sheet_name: str, records: Sequence[Mapping[str, Any]]) -> list[str]:
    if sheet_name == "Claim":
        return ["section", "field", "value"]
    names: list[str] = list(DEFAULT_COLUMNS.get(sheet_name, ()))
    seen = set(names)
    for record in records:
        for key in _flatten(record):
            if key not in seen:
                names.append(key)
                seen.add(key)
    return names or ["status"]


def _clean_text(value: str) -> str:
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Enum):
        return _excel_value(value.value)
    if isinstance(value, datetime):
        return value.astimezone(UTC).replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, (date, Decimal, int, float, bool)):
        return value
    if isinstance(value, (Mapping, list, tuple, set)):
        return _clean_text(json.dumps(json_safe(value), ensure_ascii=False, sort_keys=True))
    return _clean_text(str(value))


def _looks_money(column: str) -> bool:
    lowered = column.lower()
    return any(
        token in lowered
        for token in (
            "price",
            "amount",
            "subtotal",
            "gross",
            "vat_impact",
            "vat_total",
            "mot_total",
            "non_vat_total",
            "median_net",
            "benchmark_net",
            "invoice_net",
            "line_total",
        )
    ) and not any(token in lowered for token in ("percentage", "rate", "score", "count"))


def _write_sheet(
    workbook: Workbook,
    sheet_name: str,
    records: Sequence[Mapping[str, Any]],
    case_reference: str,
) -> None:
    ws = workbook.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True

    columns = _column_names(sheet_name, records)
    last_col = max(1, len(columns))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(1, 1, f"ClaimGuard - {sheet_name}")
    title.font = Font(name="Aptos Display", size=16, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    subtitle = ws.cell(2, 1, f"Case {case_reference or 'unassigned'} | deterministic export")
    subtitle.font = Font(name="Aptos", size=9, color=MUTED, italic=True)
    subtitle.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 18

    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(4, col_idx, name)
        cell.font = Font(name="Aptos", size=10, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        cell.border = Border(bottom=Side(style="medium", color=BLUE))
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 28

    flattened_records = [_flatten(record) for record in records]
    if not flattened_records:
        ws.cell(5, 1, "No records")
        ws.cell(5, 1).font = Font(name="Aptos", size=10, italic=True, color=MUTED)
    else:
        for row_idx, record in enumerate(flattened_records, start=5):
            for col_idx, name in enumerate(columns, start=1):
                value = _excel_value(record.get(name))
                cell = ws.cell(row_idx, col_idx, value)
                cell.font = Font(name="Aptos", size=9, color="1F2933")
                cell.alignment = Alignment(
                    horizontal=(
                        "right"
                        if isinstance(value, (Decimal, int, float)) and not isinstance(value, bool)
                        else "left"
                    ),
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = Border(bottom=THIN)
                if _looks_money(name) and isinstance(value, (Decimal, int, float)):
                    cell.number_format = "£#,##0.00;[Red]-£#,##0.00"
                elif "date" in name.lower() or name.lower().endswith("_at"):
                    if isinstance(value, datetime):
                        cell.number_format = "yyyy-mm-dd hh:mm"
                    elif isinstance(value, date):
                        cell.number_format = "yyyy-mm-dd"

        table_ref = f"A4:{get_column_letter(last_col)}{4 + len(flattened_records)}"
        table = Table(displayName=f"ClaimGuard{sheet_name}Table", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.auto_filter.ref = f"A4:{get_column_letter(last_col)}{max(5, 4 + len(flattened_records))}"
    ws.print_title_rows = "1:4"
    ws.print_area = f"A1:{get_column_letter(last_col)}{max(5, 4 + len(flattened_records))}"

    for col_idx, name in enumerate(columns, start=1):
        values = [str(name)]
        values.extend(str(record.get(name, "")) for record in flattened_records[:250])
        width = min(48, max(11, max((len(value) for value in values), default=11) + 2))
        if any(
            token in name.lower()
            for token in ("description", "reason", "rationale", "message", "summary", "note")
        ):
            width = min(48, max(width, 28))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if sheet_name == "Challenges" and flattened_records:
        included_col = (
            columns.index("included_in_letter") + 1 if "included_in_letter" in columns else None
        )
        amount_col = (
            columns.index("computed_challenge_amount_net") + 1
            if "computed_challenge_amount_net" in columns
            else None
        )
        if included_col and amount_col:
            data_range = f"A5:{get_column_letter(last_col)}{4 + len(flattened_records)}"
            included_letter = f"${get_column_letter(included_col)}5=TRUE"
            excluded_letter = f"${get_column_letter(included_col)}5=FALSE"
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(formula=[included_letter], fill=PatternFill("solid", fgColor=PALE_RED)),
            )
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[excluded_letter], fill=PatternFill("solid", fgColor=PALE_GREEN)
                ),
            )


def build_case_workbook(result: Mapping[str, Any], output_path: str | Path) -> Path:
    """Create the twelve-sheet per-case audit workbook with typed values."""

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "ClaimGuard"
    workbook.properties.title = "ClaimGuard case audit export"
    workbook.properties.subject = "Motor claims invoice validation and challenge evidence"
    workbook.properties.keywords = "ClaimGuard, motor claim, invoice, audit"

    claim = case_record(result)
    case_reference = str(
        first_value(claim, ("case_reference", "claim_number", "reference", "id"), "")
    )
    for sheet_name in SHEET_ORDER:
        _write_sheet(
            workbook,
            sheet_name,
            _sheet_records(result, sheet_name),
            case_reference,
        )

    with NamedTemporaryFile(
        prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent, delete=False
    ) as handle:
        temp_path = Path(handle.name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    finally:
        temp_path.unlink(missing_ok=True)
    return path
