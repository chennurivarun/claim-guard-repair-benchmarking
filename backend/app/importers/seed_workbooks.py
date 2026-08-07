from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    return Decimal(str(value))


def _date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def _tokens(value: Any) -> tuple[str, ...]:
    if value is None:
        return ()
    return tuple(token.strip() for token in str(value).split(";") if token.strip())


def _rows(path: Path, sheet_name: str) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Workbook {path.name} is missing sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = tuple(str(value).strip() for value in next(iterator))
    try:
        for values in iterator:
            if not any(value is not None and value != "" for value in values):
                continue
            yield dict(zip(headers, values, strict=False))
    finally:
        workbook.close()


@dataclass(frozen=True)
class OntologySeedRecord:
    ontology_id: str
    canonical_name: str
    item_type: str
    category: str
    unit: str
    reference_price_net: Decimal | None
    price_vat_basis: str
    currency: str
    synonyms: tuple[str, ...]
    part_number_examples: tuple[str, ...]
    part_grade: str | None
    vehicle_applicability: str | None
    region: str | None
    price_source: str
    observation_count: int
    effective_date: date | None
    approval_status: str
    confidence_level: str | None
    notes: str | None


@dataclass(frozen=True)
class HistoricalSeedRecord:
    claim_line_id: str
    source_file: str
    invoice_number: str
    invoice_date: date
    garage_name: str
    workshop_category: str | None
    region: str | None
    vehicle_make: str | None
    vehicle_model: str | None
    vehicle_year: int | None
    official_vehicle_class: str | None
    bodywork_code: str | None
    market_segment: str | None
    classification_source: str | None
    registration: str | None
    mileage: int | None
    document_role: str
    item_kind: str
    raw_description: str
    part_number: str | None
    quantity: Decimal | None
    unit_price_net: Decimal | None
    line_total_net: Decimal
    vat_rate: Decimal
    mapped_ontology_id: str
    notes: str | None


@dataclass(frozen=True)
class InvoiceSummarySeedRecord:
    invoice_number: str
    invoice_date: date
    garage_name: str
    document_role: str
    line_count: int
    sum_lines_net: Decimal
    gross_total: Decimal | None


@dataclass(frozen=True)
class SeedWorkbookBundle:
    ontology_items: tuple[OntologySeedRecord, ...]
    runtime_history: tuple[HistoricalSeedRecord, ...]
    invoice_summaries: tuple[InvoiceSummarySeedRecord, ...]
    acceptance_gold: tuple[HistoricalSeedRecord, ...]


def _ontology_record(row: dict[str, Any]) -> OntologySeedRecord:
    price = _decimal(row["reference_price_gbp_net"])
    return OntologySeedRecord(
        ontology_id=str(row["ontology_id"]).strip(),
        canonical_name=str(row["canonical_name"]).strip(),
        item_type=str(row["item_type"]).strip(),
        category=str(row["category"]).strip(),
        unit=str(row["unit"]).strip().lower(),
        reference_price_net=price,
        price_vat_basis=str(row["price_vat_basis"]).strip(),
        currency=str(row["currency"]).strip(),
        synonyms=_tokens(row.get("synonyms")),
        part_number_examples=_tokens(row.get("part_number_examples")),
        part_grade=str(row["part_grade"]).strip() if row.get("part_grade") else None,
        vehicle_applicability=(
            str(row["vehicle_applicability"]).strip() if row.get("vehicle_applicability") else None
        ),
        region=str(row["region"]).strip() if row.get("region") else None,
        price_source=str(row["price_source"]).strip(),
        observation_count=int(row.get("n_observations") or 0),
        effective_date=_date(row.get("effective_date")),
        approval_status=str(row["approval_status"]).strip(),
        confidence_level=(
            str(row["confidence_level"]).strip() if row.get("confidence_level") else None
        ),
        notes=str(row["notes"]).strip() if row.get("notes") else None,
    )


def _history_record(row: dict[str, Any]) -> HistoricalSeedRecord:
    invoice_date = _date(row.get("invoice_date"))
    line_total = _decimal(row.get("line_total_net_gbp"))
    if invoice_date is None or line_total is None:
        raise ValueError(f"History line {row.get('claim_line_id')} lacks date or line total")
    return HistoricalSeedRecord(
        claim_line_id=str(row["claim_line_id"]).strip(),
        source_file=str(row["source_file"]).strip(),
        invoice_number=str(row["invoice_number"]).strip(),
        invoice_date=invoice_date,
        garage_name=str(row["garage_name"]).strip(),
        workshop_category=(
            str(row["workshop_category"]).strip() if row.get("workshop_category") else None
        ),
        region=str(row["region"]).strip() if row.get("region") else None,
        vehicle_make=str(row["vehicle_make"]).strip() if row.get("vehicle_make") else None,
        vehicle_model=str(row["vehicle_model"]).strip() if row.get("vehicle_model") else None,
        vehicle_year=int(row["vehicle_year"]) if row.get("vehicle_year") else None,
        official_vehicle_class=(
            str(row["official_vehicle_class"]).strip()
            if row.get("official_vehicle_class")
            else None
        ),
        bodywork_code=str(row["bodywork_code"]).strip() if row.get("bodywork_code") else None,
        market_segment=(str(row["market_segment"]).strip() if row.get("market_segment") else None),
        classification_source=(
            str(row["classification_source"]).strip() if row.get("classification_source") else None
        ),
        registration=str(row["vrm"]).strip() if row.get("vrm") else None,
        mileage=int(row["mileage"]) if row.get("mileage") else None,
        document_role=str(row["document_role"]).strip().lower(),
        item_kind=str(row["item_kind"]).strip().lower(),
        raw_description=str(row["raw_description"]).strip(),
        part_number=str(row["part_number"]).strip() if row.get("part_number") else None,
        quantity=_decimal(row.get("quantity")),
        unit_price_net=_decimal(row.get("unit_price_net_gbp")),
        line_total_net=line_total,
        vat_rate=_decimal(row.get("vat_rate_pct")) or Decimal("0"),
        mapped_ontology_id=str(row["mapped_ontology_id"]).strip(),
        notes=str(row["notes"]).strip() if row.get("notes") else None,
    )


def _summary_record(row: dict[str, Any]) -> InvoiceSummarySeedRecord:
    invoice_date = _date(row.get("invoice_date"))
    sum_lines = _decimal(row.get("sum_of_lines_net_gbp"))
    if invoice_date is None or sum_lines is None:
        raise ValueError(f"Invoice summary {row.get('invoice_number')} lacks date or total")
    return InvoiceSummarySeedRecord(
        invoice_number=str(row["invoice_number"]).strip(),
        invoice_date=invoice_date,
        garage_name=str(row["garage_name"]).strip(),
        document_role=str(row["document_role"]).strip().lower(),
        line_count=int(row["n_lines"]),
        sum_lines_net=sum_lines,
        gross_total=_decimal(row.get("gross_total_as_shown_gbp")),
    )


def load_seed_workbooks(
    ontology_path: str | Path,
    historical_path: str | Path,
) -> SeedWorkbookBundle:
    """Load seed banks while keeping current test invoices out of runtime history."""

    ontology_path = Path(ontology_path)
    historical_path = Path(historical_path)
    ontology = tuple(_ontology_record(row) for row in _rows(ontology_path, "ontology_items"))
    runtime_history = tuple(
        _history_record(row) for row in _rows(historical_path, "claims_line_items")
    )
    summaries = tuple(_summary_record(row) for row in _rows(historical_path, "invoice_summary"))
    acceptance_gold = tuple(
        _history_record(row) for row in _rows(historical_path, "current_test_invoices")
    )
    return SeedWorkbookBundle(
        ontology_items=ontology,
        runtime_history=runtime_history,
        invoice_summaries=summaries,
        acceptance_gold=acceptance_gold,
    )
